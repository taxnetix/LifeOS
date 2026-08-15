"""XLSX extraction — broker and platform statements.

Two decisions worth stating:

  * `data_only=True` reads cached formula RESULTS, not formulas. A ledger needs
    the number the human saw. If a file was written by a tool that never cached
    results, the cells come back None — reported as an error rather than
    silently becoming an empty sheet.

  * The locator is the real cell reference (`sheet=Summary;cell=B14`), so a
    figure in a report can be pointed at precisely.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from . import Block, Extraction, Table

_CONF = 0.98  # values are exact; only header/shape interpretation can mislead
_MAX_ROWS = 20_000


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_xlsx(path: Path, doc_hash: str) -> Extraction:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    out = Extraction(doc_hash=doc_hash, method="openpyxl", pages=0)
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    out.pages = len(wb.sheetnames)
    out.meta = {"sheets": ", ".join(wb.sheetnames)}
    all_none = True

    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[list[str]] = []
            for r, row in enumerate(ws.iter_rows(values_only=False), 1):
                if r > _MAX_ROWS:
                    out.errors.append(f"sheet '{name}' truncated at {_MAX_ROWS} rows")
                    break
                values = []
                for cell in row:
                    if cell.value is not None:
                        all_none = False
                        out.blocks.append(
                            Block(
                                locator=f"sheet={name};cell={get_column_letter(cell.column)}{cell.row}",
                                text=_cell(cell.value),
                                confidence=_CONF,
                            )
                        )
                    values.append(_cell(cell.value))
                if any(values):
                    rows.append(values)
            if rows:
                out.tables.append(Table(locator=f"sheet={name}", rows=rows, confidence=_CONF))
    finally:
        wb.close()

    if all_none and out.tables:
        out.errors.append(
            "every cell is empty — the workbook may store formulas with no cached "
            "results; open and re-save it in Excel, or export to CSV"
        )
    if not out.tables:
        out.errors.append("workbook contains no data")
    return out
